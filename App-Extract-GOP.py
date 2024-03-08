import tarfile
import pandas as pd
import os
import tkinter as tk
from tkinter import filedialog, messagebox
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment
from decimal import Decimal, ROUND_DOWN
import re

# Placeholder for combined Thermal Data
all_thermal_df = pd.DataFrame()

# Function to return one digit after decimal
def keep_first_digit_after_decimal(num):
    # Convert the number to a Decimal
    decimal_num = Decimal(str(num))
    # Use the quantize method to keep only one digit after the decimal point without rounding
    new_num = decimal_num.quantize(Decimal('0.0'), rounding=ROUND_DOWN)
    return float(new_num)

# Function to return one digit after decimal
def keep_two_digit_after_decimal(num):
    # Convert the number to a Decimal
    decimal_num = Decimal(str(num))
    # Use the quantize method to keep only one digit after the decimal point without rounding
    new_num = decimal_num.quantize(Decimal('0.00'), rounding=ROUND_DOWN)
    return float(new_num)

# Function to process a single tarball
def process_tarball(tarball_path, combined_data, combined_thermal_data, combined_rbf_data, tarball_name):
    with tarfile.open(tarball_path, 'r') as tar:
        print(f"Extracting {tarball_path}...")
        # Extract the data-picture-analysis.txt file
        tar.extract('data-picture-analysis.txt')
        # Extract the thermal_log.csv file
        tar.extract('thermal_log.csv')
        # Extract the custom-runscript.py file
        tar.extract('custom-runscript.py')

    variable_name_PCR_mapping = {
    "EXPERIMENT_ID": "Experiment ID",
    "RUN_DATE": "Run Date",
    "OPERATOR_NAME": "Operator Name",
    "NOTES": "Notes",
    "AMPLIFICATION_METHOD": "Amplification Method",
    "NUMBER_OF_CYCLES": "Number of Cycles",
    "INCUBATION_1_TARGET_TEMP": "Incubation 1 Target Sleeve Temperature in °C",
    "INCUBATION_1_RAMP_RATE": "Incubation 1 Ramp Rate in °C/Second",
    "INCUBATION_1_DWELL_TIME": "Incubation 1 Dwell Time in Seconds",
    "INCUBATION_2_TARGET_TEMP": "Incubation 2 Target Sleeve Temperature in °C",
    "INCUBATION_2_RAMP_RATE": "Incubation 2 Ramp Rate in °C/Second",
    "INCUBATION_2_DWELL_TIME": "Incubation 2 Dwell Time in Seconds",
    "INCUBATION_3_TARGET_TEMP": "Incubation 3 Target Sleeve Temperature in °C",
    "INCUBATION_3_RAMP_RATE": "Incubation 3 Ramp Rate in °C/Second",
    "INCUBATION_3_DWELL_TIME": "Incubation 3 Dwell Time in Seconds",
    "CYCLIC_DENATURE_TEMP": "Denaturation Temperature in °C",
    "CYCLIC_DENATURE_RAMP_RATE": "Denaturation Ramp Rate  in °C/Second",
    "CYCLIC_DENATURE_DWELL_TIME": "Denaturation Dwell Time in Seconds",
    "CYCLIC_ANNEAL_TEMP": "Annealing Temperature in °C",
    "CYCLIC_ANNEAL_RAMP_RATE": "Annealing Ramp Rate °C/Second",
    "CYCLIC_ANNEAL_DWELL_TIME": "Annealing Dwell Time in Seconds"
   }

    variable_name_ISO_mapping = {
    "EXPERIMENT_ID": "Experiment ID",
    "RUN_DATE": "Run Date",
    "OPERATOR_NAME": "Operator Name",
    "NOTES": "Notes",
    "AMPLIFICATION_METHOD": "Amplification Method",
    "INCUBATION_1_TARGET_TEMP": "Incubation 1 Target Sleeve Temperature in °C",
    "INCUBATION_1_RAMP_RATE": "Incubation 1 Ramp Rate in °C/Second",
    "INCUBATION_1_DWELL_TIME": "Incubation 1 Dwell Time in Seconds",
    "INCUBATION_2_TARGET_TEMP": "Incubation 2 Target Sleeve Temperature in °C",
    "INCUBATION_2_RAMP_RATE": "Incubation 2 Ramp Rate in °C/Second",
    "INCUBATION_2_DWELL_TIME": "Incubation 2 Dwell Time in Seconds",
    "INCUBATION_3_TARGET_TEMP": "Incubation 3 Target Sleeve Temperature in °C",
    "INCUBATION_3_RAMP_RATE": "Incubation 3 Ramp Rate in °C/Second",
    "INCUBATION_3_DWELL_TIME": "Incubation 3 Dwell Time in Seconds",
    "IMAGE_CAPTURE_PERIOD": "Image Capture Frequency in Seconds",
    "ISOTHERMAL_AMPLIFICATION_TEMP": "Isothermal Target Sleeve Temperature in °C",
    "ISOTHERMAL_AMPLIFICATION_TIME": "Isothermal Target Sleeve Time in Seconds"
   }

    data = []

    # Getting Experiment ID
    with open('custom-runscript.py', 'r') as file:
        script_content = file.read()
        id_var = "EXPERIMENT_ID"
        if id_var in script_content:
            # Use a regular expression to find the assignment line for the variable
            pattern = f"{id_var} = (.*?)\\n"
            match = re.search(pattern, script_content, re.DOTALL)
            if match:
              id_var_value = match.group(1).strip('"')

    # Getting Fluorescence Data
    with open('data-picture-analysis.txt', 'r') as file:
        for i, line in enumerate(file, start=1):
            values = line.strip().split()
            values = [float(value) for value in values]
            modified_numbers = [keep_first_digit_after_decimal(num) for num in values]
            data.append([id_var_value,tarball_name[0:9], tarball_name, i] + modified_numbers)

    combined_data.extend(data)

    # Getting Thermal Data
    thermal_data = []
    with open('thermal_log.csv', 'r') as file:
        for line in file:
            values = line.split(',')
            modified_values = []

            for i, value in enumerate(values):
                value = value.strip()
                if value.replace('.', '').isdigit():
                    if i == 1:  # Check if it's the second element (0-based index)
                        modified_values.append(keep_first_digit_after_decimal(float(value)))
                    else:
                        modified_values.append(keep_two_digit_after_decimal(float(value)))
                else:
                    modified_values.append(value.strip())
            thermal_data.append(modified_values[:5])

    combined_thermal_data.extend(thermal_data)

    # Getting RBF Data
    rbf_data = []
    with open('custom-runscript.py', 'r') as file:
        script_content = file.read()
        amp_var = "AMPLIFICATION_METHOD"
        if amp_var in script_content:
            amp_pattern =  f"{amp_var} = (.*?)\\n"
            amp_match = re.search(amp_pattern, script_content, re.DOTALL)
            if amp_match:
                amp_var_value = amp_match.group(1).strip('"')

        if amp_var_value == "PCR":
            for original_var_name, new_var_name in variable_name_PCR_mapping.items():
                # Check if the original variable name exists in the custom runscript
                if original_var_name in script_content:
                    # Use a regular expression to find the assignment line for the variable
                    pattern = f"{original_var_name} = (.*?)\\n"
                    match = re.search(pattern, script_content, re.DOTALL)
                    if match:
                        # Extract the variable value and format the line
                        var_value = match.group(1).strip('"')
                        rbf_data.append(var_value)
        if amp_var_value == "ISOTHERMAL":
            for original_var_name, new_var_name in variable_name_ISO_mapping.items():
                # Check if the original variable name exists in the custom runscript
                if original_var_name in script_content:
                    # Use a regular expression to find the assignment line for the variable
                    pattern = f"{original_var_name} = (.*?)\\n"
                    match = re.search(pattern, script_content, re.DOTALL)
                    if match:
                        # Extract the variable value and format the line
                        var_value = match.group(1).strip('"')
                        rbf_data.append(var_value)

    combined_rbf_data.extend(rbf_data)

    # Clean up the extracted files
    os.remove('data-picture-analysis.txt')
    os.remove('thermal_log.csv')
    os.remove('custom-runscript.py')

def browse_tarball_directory():
    folder_path = filedialog.askdirectory()
    tarball_directory_entry.delete(0, tk.END)
    tarball_directory_entry.insert(0, folder_path)

def browse_output_directory():
    folder_path = filedialog.askdirectory()
    output_directory_entry.delete(0, tk.END)
    output_directory_entry.insert(0, folder_path)

def process_tarballs():
    tarball_dir = tarball_directory_entry.get()
    output_dir = output_directory_entry.get()

    if not tarball_dir or not output_dir:
        messagebox.showerror("Error", "One or more of the extractor requirements are missing. Please ensure you specify both, the current location of the tarball directory and the desired location for the output file.")
        return

    combined_data = []
    combined_thermal_data = []
    combined_rbf_data = []

    wb = Workbook()
    # Remove the default "Sheet"
    default_sheet = wb['Sheet']
    wb.remove(default_sheet)
    picture_analysis_sheet = wb.create_sheet("Fluorescence Data")
    header_sheet = list(range(1, 121))
    header_sheet.insert(0, "Channel")
    header_sheet.insert(0, "Experiment ID")
    header_sheet.insert(0, 'Tarball ID')
    picture_analysis_sheet.append(header_sheet)

    thermal_log_sheet = wb.create_sheet("Thermal Data")

    rbf_sheet = wb.create_sheet("Run Information (RBF Data)", 0)
    rbf_headers = ["Tarball ID", "Cube ID", "Experiment ID", "Run Date", "Operator Name", "Notes", "Amplification Method", "Number of Cycles",
                    "Incubation 1 Target Sleeve Temperature in °C", "Incubation 1 Ramp Rate in °C/Second", "Incubation 1 Dwell Time in Seconds",
                   "Incubation 2 Target Sleeve Temperature in °C", "Incubation 2 Ramp Rate in °C/Second", "Incubation 2 Dwell Time in Seconds",
                   "Incubation 3 Target Sleeve Temperature in °C", "Incubation 3 Ramp Rate in °C/Second", "Incubation 3 Dwell Time in Seconds", 
                   "Denaturation Temperature in °C", "Denaturation Ramp Rate in °C/Second", "Denaturation Dwell Time in Seconds", 
                   "Annealing Temperature in °C", "Annealing Ramp Rate in °C/Second", "Annealing Dwell Time in Seconds",
                   "Isothermal Image Capture Frequency in Seconds", "Isothermal Target Sleeve Temeperature in °C", "Isothermal Dwell Time in Seconds"]
    rbf_sheet.append(rbf_headers)

    # Loop through the tarballs and process each one
    for filename in os.listdir(tarball_dir):
        if filename.endswith('.tar.gz'):
            tarball_path = os.path.join(tarball_dir, filename)
            tarball_name = os.path.splitext(filename)[0]
            process_tarball(tarball_path, combined_data, combined_thermal_data, combined_rbf_data, tarball_name)

            # Fluorescence Tab
            r1 = [x[4] for x in combined_data]
            r2 = [x[5] for x in combined_data]
            r3 = [x[6] for x in combined_data]
            g1 = [x[7] for x in combined_data]
            g2 = [x[8] for x in combined_data]
            g3 = [x[9] for x in combined_data]
            g1.insert(0, "Green-1")
            g1.insert(0, combined_data[0][0])
            g1.insert(0, combined_data[0][2])
            r1.insert(0, "Red-1")
            r1.insert(0, combined_data[0][0])
            r1.insert(0, combined_data[0][2])
            g2.insert(0, "Green-2")
            g2.insert(0, combined_data[0][0])
            g2.insert(0, combined_data[0][2])
            r2.insert(0, "Red-2")
            r2.insert(0, combined_data[0][0])
            r2.insert(0, combined_data[0][2])
            g3.insert(0, "Green-3")
            g3.insert(0, combined_data[0][0])
            g3.insert(0, combined_data[0][2])
            r3.insert(0, "Red-3")
            r3.insert(0, combined_data[0][0])
            r3.insert(0, combined_data[0][2])
            picture_analysis_sheet.append(g1)
            picture_analysis_sheet.append(r1)
            picture_analysis_sheet.append(g2)
            picture_analysis_sheet.append(r2)
            picture_analysis_sheet.append(g3)
            picture_analysis_sheet.append(r3)
            picture_analysis_sheet.append(['End of this Tarball'])
            
            # Thermal Tab
            thermal_header = ["Seconds elapsed from test start (Secs)", "Actual Sleeve Temperature (°C)", "Target Sleeve Temperature (°C)", "Thermal-System", "Cycle Number"]
            thermal_df = pd.DataFrame(combined_thermal_data, columns=thermal_header)
            # Strip whitespace and convert to lowercase for comparison
            thermal_df = thermal_df[thermal_df['Thermal-System'].str.strip().str.lower() == 'sleeve']
            thermal_df = thermal_df.drop(["Thermal-System"], axis=1)
            thermal_df = thermal_df.drop(["Target Sleeve Temperature (°C)"], axis=1)
            thermal_df.loc[-1] = [combined_data[0][0], combined_data[0][0], combined_data[0][0]]
            thermal_df.index = thermal_df.index + 1
            thermal_df = thermal_df.sort_index().reset_index(drop=True)
            thermal_df.loc[-1] = [combined_data[0][2], combined_data[0][2], combined_data[0][2]]
            thermal_df.index = thermal_df.index + 1
            thermal_df = thermal_df.sort_index().reset_index(drop=True)
            thermal_df["End of this Tarball"] = ""
            global all_thermal_df
            all_thermal_df = pd.concat([all_thermal_df, thermal_df], axis=1)

            # RBF Tab
            if combined_rbf_data[4] == "PCR":
                tid = tarball_name
                if tid[0]=='l':
                    tarball_name = tarball_name[6:]
                cid = tarball_name[0:9]
                combined_rbf_data.insert(0, cid)
                combined_rbf_data.insert(0, tid)
                combined_rbf_data.append("N/A")
                combined_rbf_data.append("N/A")
                combined_rbf_data.append("N/A")
                rbf_sheet.append(combined_rbf_data)
            if combined_rbf_data[4] == "ISOTHERMAL":
                tid = tarball_name
                if tid[0]=='l':
                    tarball_name = tarball_name[6:]
                cid = tarball_name[0:9]
                combined_rbf_data.insert(0, cid)
                combined_rbf_data.insert(0, tid)
                combined_rbf_data.insert(7, "N/A")
                combined_rbf_data.insert(17, "N/A")
                combined_rbf_data.insert(18, "N/A")
                combined_rbf_data.insert(19, "N/A")
                combined_rbf_data.insert(20, "N/A")
                combined_rbf_data.insert(21, "N/A")
                combined_rbf_data.insert(22, "N/A")
                rbf_sheet.append(combined_rbf_data)

            combined_data = []
            combined_thermal_data = []
            combined_rbf_data = []

    #Bold Headers
    header_font = Font(bold=True)
    header_alignment = Alignment(horizontal='center', vertical='center')
    for cell in picture_analysis_sheet['1']:
        cell.font = header_font
        cell.alignment = header_alignment

    # Autofit column width of the Excel
    for column_letter in picture_analysis_sheet.columns:
        max_length = 0
        for cell in column_letter:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        picture_analysis_sheet.column_dimensions[column_letter[0].column_letter].width = adjusted_width

    for row in dataframe_to_rows(all_thermal_df, index=False, header=True):
        thermal_log_sheet.append(row)
    
    # Bold Headers
    for cell in thermal_log_sheet['1']:
        cell.font = header_font
        cell.alignment = header_alignment

    for cell in thermal_log_sheet['2']:
        cell.alignment = header_alignment

    for cell in thermal_log_sheet['3']:
        cell.alignment = header_alignment

    # Autofit column width of the Excel
    for column_letter in thermal_log_sheet.columns:
        max_length = 0
        for cell in column_letter:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        thermal_log_sheet.column_dimensions[column_letter[0].column_letter].width = adjusted_width

    #Bold Headers
    for cell in rbf_sheet['1']:
        cell.font = header_font
        cell.alignment = header_alignment

    # Autofit column width of the Excel
    for column_letter in rbf_sheet.columns:
        max_length = 0
        for cell in column_letter:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        rbf_sheet.column_dimensions[column_letter[0].column_letter].width = adjusted_width

    # Save the workbook
    output_excel = os.path.join(output_dir, f"{output_excel_name_entry.get()}.xlsx")
    if not output_excel_name_entry.get():
        messagebox.showerror("Error", "Please type the name of the Excel file.")
        return
    wb.save(output_excel)

    print(f"Saved {output_excel}")

    # Show a "Done" message
    messagebox.showinfo("Done", "All tarballs processed and saved as a single Excel file.")

# Create the main tkinter window
window = tk.Tk()
window.title("Genomadix CubeX Tarball Extractor")

window.geometry("500x300")

# Tarball Directory Selection
tarball_directory_label = tk.Label(window, text="Select Tarball Directory:")
tarball_directory_label.pack()

tarball_directory_entry = tk.Entry(window, width=50)
tarball_directory_entry.pack()

browse_tarball_button = tk.Button(window, text="Browse", command=browse_tarball_directory)
browse_tarball_button.pack()

# Output Directory Selection
output_directory_label = tk.Label(window, text="Select Output Directory:")
output_directory_label.pack()

output_directory_entry = tk.Entry(window, width=50)
output_directory_entry.pack()

browse_output_button = tk.Button(window, text="Browse", command=browse_output_directory)
browse_output_button.pack()

# Output Excel Name Entry
output_excel_name_label = tk.Label(window, text="Output Excel Name:")
output_excel_name_label.pack()

output_excel_name_entry = tk.Entry(window, width=50)
output_excel_name_entry.pack()

# Process Tarballs Button
process_button = tk.Button(window, text="Process Tarballs", command=process_tarballs)
process_button.pack()

# Start the tkinter event loop
window.mainloop()