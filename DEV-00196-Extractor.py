import os
import glob
import tarfile
import gzip
import json
import pandas as pd
import openpyxl
import json
import tkinter as tk
from tkinter import filedialog

def process_tar_file(tar_file, output_df, row_index, wb):

    second_sheet = wb["optical_store"]
    third_sheet = wb["optical_targets"]

    with tarfile.open(tar_file, "r:*") as tar:
        hostname_data = None
        ref_unit_data = None
        batch_code_data = None
        date_time_data = None
        red_16x_w1 = None
        red_16x_w2 = None
        red_16x_w3 = None
        
        for member in tar.getmembers():
            if member.name.endswith("device_information.json"):
                f = tar.extractfile(member)
                json_data = json.load(f)
                hostname_data = json_data["hostname"]
                second_sheet.cell(row=1, column=3).value = hostname_data
            elif member.name.endswith("optical_cal_targets_gen_3_0.json"):
                f = tar.extractfile(member)
                json_data = json.load(f)

                # data_targets = json.load(f)
                flattened_data_targets = flatten_dict(json_data)
                idx = 1
                for key, value in flattened_data_targets.items():
                    third_sheet.cell(row=idx, column=1).value = key
                    third_sheet.cell(row=idx, column=2).value = value
                    idx += 1

                ref_unit_data = json_data["ref_unit"]
                batch_code_data = json_data["batchcode"]
                date_time_data = json_data["timestamp"]
            elif member.name.endswith("well_center_store.json"):
                f = tar.extractfile(member)
                json_data = json.load(f)
                well_1_center = json_data['test_results']['wells']['well_1']
                well_2_center = json_data['test_results']['wells']['well_2']
                well_3_center = json_data['test_results']['wells']['well_3']
            elif member.name.endswith("optical_store_gen3_storage.json"):
                f = tar.extractfile(member)
                json_data = json.load(f)

                # data_store = json.load(f)
                flattened_data_store = flatten_dict(json_data)
                idx = 1
                for key, value in flattened_data_store.items():
                    second_sheet.cell(row=idx, column=1).value = key
                    second_sheet.cell(row=idx, column=2).value = value
                    idx += 1

                #Red Standards 1,2,3
                red_16x_w1 = json_data['test_results']['analysis_values']['postcal']['standard_1']['img1'][0]
                red_16x_w2 = json_data['test_results']['analysis_values']['postcal']['standard_1']['img1'][1]
                red_16x_w3 = json_data['test_results']['analysis_values']['postcal']['standard_1']['img1'][2]
                red_16x_w1_p2 = json_data['test_results']['analysis_values']['postcal']['standard_1']['img2'][0]
                red_16x_w2_p2 = json_data['test_results']['analysis_values']['postcal']['standard_1']['img2'][1]
                red_16x_w3_p2 = json_data['test_results']['analysis_values']['postcal']['standard_1']['img2'][2]
                red_16x_w1_p3 = json_data['test_results']['analysis_values']['postcal']['standard_1']['img3'][0]
                red_16x_w2_p3 = json_data['test_results']['analysis_values']['postcal']['standard_1']['img3'][1]
                red_16x_w3_p3 = json_data['test_results']['analysis_values']['postcal']['standard_1']['img3'][2]

                red_32x_w1 = json_data['test_results']['analysis_values']['postcal']['standard_2']['img1'][0]
                red_32x_w2 = json_data['test_results']['analysis_values']['postcal']['standard_2']['img1'][1]
                red_32x_w3 = json_data['test_results']['analysis_values']['postcal']['standard_2']['img1'][2]
                red_32x_w1_p2 = json_data['test_results']['analysis_values']['postcal']['standard_2']['img2'][0]
                red_32x_w2_p2 = json_data['test_results']['analysis_values']['postcal']['standard_2']['img2'][1]
                red_32x_w3_p2 = json_data['test_results']['analysis_values']['postcal']['standard_2']['img2'][2]
                red_32x_w1_p3 = json_data['test_results']['analysis_values']['postcal']['standard_2']['img3'][0]
                red_32x_w2_p3 = json_data['test_results']['analysis_values']['postcal']['standard_2']['img3'][1]
                red_32x_w3_p3 = json_data['test_results']['analysis_values']['postcal']['standard_2']['img3'][2]

                red_64x_w1 = json_data['test_results']['analysis_values']['postcal']['standard_3']['img1'][0]
                red_64x_w2 = json_data['test_results']['analysis_values']['postcal']['standard_3']['img1'][1]
                red_64x_w3 = json_data['test_results']['analysis_values']['postcal']['standard_3']['img1'][2]
                red_64x_w1_p2 = json_data['test_results']['analysis_values']['postcal']['standard_3']['img2'][0]
                red_64x_w2_p2 = json_data['test_results']['analysis_values']['postcal']['standard_3']['img2'][1]
                red_64x_w3_p2 = json_data['test_results']['analysis_values']['postcal']['standard_3']['img2'][2]
                red_64x_w1_p3 = json_data['test_results']['analysis_values']['postcal']['standard_3']['img3'][0]
                red_64x_w2_p3 = json_data['test_results']['analysis_values']['postcal']['standard_3']['img3'][1]
                red_64x_w3_p3 = json_data['test_results']['analysis_values']['postcal']['standard_3']['img3'][2]

                #Red Standards 4,5,6
                red_16x_w1_std4 = json_data['test_results']['analysis_values']['postcal']['standard_4']['img1'][0]
                red_16x_w2_std4 = json_data['test_results']['analysis_values']['postcal']['standard_4']['img1'][1]
                red_16x_w3_std4 = json_data['test_results']['analysis_values']['postcal']['standard_4']['img1'][2]
                red_16x_w1_p2_std4 = json_data['test_results']['analysis_values']['postcal']['standard_4']['img2'][0]
                red_16x_w2_p2_std4 = json_data['test_results']['analysis_values']['postcal']['standard_4']['img2'][1]
                red_16x_w3_p2_std4 = json_data['test_results']['analysis_values']['postcal']['standard_4']['img2'][2]
                red_16x_w1_p3_std4 = json_data['test_results']['analysis_values']['postcal']['standard_4']['img3'][0]
                red_16x_w2_p3_std4 = json_data['test_results']['analysis_values']['postcal']['standard_4']['img3'][1]
                red_16x_w3_p3_std4 = json_data['test_results']['analysis_values']['postcal']['standard_4']['img3'][2]

                red_32x_w1_std5 = json_data['test_results']['analysis_values']['postcal']['standard_5']['img1'][0]
                red_32x_w2_std5 = json_data['test_results']['analysis_values']['postcal']['standard_5']['img1'][1]
                red_32x_w3_std5 = json_data['test_results']['analysis_values']['postcal']['standard_5']['img1'][2]
                red_32x_w1_p2_std5 = json_data['test_results']['analysis_values']['postcal']['standard_5']['img2'][0]
                red_32x_w2_p2_std5 = json_data['test_results']['analysis_values']['postcal']['standard_5']['img2'][1]
                red_32x_w3_p2_std5 = json_data['test_results']['analysis_values']['postcal']['standard_5']['img2'][2]
                red_32x_w1_p3_std5 = json_data['test_results']['analysis_values']['postcal']['standard_5']['img3'][0]
                red_32x_w2_p3_std5 = json_data['test_results']['analysis_values']['postcal']['standard_5']['img3'][1]
                red_32x_w3_p3_std5 = json_data['test_results']['analysis_values']['postcal']['standard_5']['img3'][2]

                red_64x_w1_std6 = json_data['test_results']['analysis_values']['postcal']['standard_6']['img1'][0]
                red_64x_w2_std6 = json_data['test_results']['analysis_values']['postcal']['standard_6']['img1'][1]
                red_64x_w3_std6 = json_data['test_results']['analysis_values']['postcal']['standard_6']['img1'][2]
                red_64x_w1_p2_std6 = json_data['test_results']['analysis_values']['postcal']['standard_6']['img2'][0]
                red_64x_w2_p2_std6 = json_data['test_results']['analysis_values']['postcal']['standard_6']['img2'][1]
                red_64x_w3_p2_std6 = json_data['test_results']['analysis_values']['postcal']['standard_6']['img2'][2]
                red_64x_w1_p3_std6 = json_data['test_results']['analysis_values']['postcal']['standard_6']['img3'][0]
                red_64x_w2_p3_std6 = json_data['test_results']['analysis_values']['postcal']['standard_6']['img3'][1]
                red_64x_w3_p3_std6 = json_data['test_results']['analysis_values']['postcal']['standard_6']['img3'][2]

                #Green Standards 1,2,3
                green_16x_w1 = json_data['test_results']['analysis_values']['postcal']['standard_1']['img1'][3]
                green_16x_w2 = json_data['test_results']['analysis_values']['postcal']['standard_1']['img1'][4]
                green_16x_w3 = json_data['test_results']['analysis_values']['postcal']['standard_1']['img1'][5]
                green_16x_w1_p2 = json_data['test_results']['analysis_values']['postcal']['standard_1']['img2'][3]
                green_16x_w2_p2 = json_data['test_results']['analysis_values']['postcal']['standard_1']['img2'][4]
                green_16x_w3_p2 = json_data['test_results']['analysis_values']['postcal']['standard_1']['img2'][5]
                green_16x_w1_p3 = json_data['test_results']['analysis_values']['postcal']['standard_1']['img3'][3]
                green_16x_w2_p3 = json_data['test_results']['analysis_values']['postcal']['standard_1']['img3'][4]
                green_16x_w3_p3 = json_data['test_results']['analysis_values']['postcal']['standard_1']['img3'][5]

                green_32x_w1 = json_data['test_results']['analysis_values']['postcal']['standard_2']['img1'][3]
                green_32x_w2 = json_data['test_results']['analysis_values']['postcal']['standard_2']['img1'][4]
                green_32x_w3 = json_data['test_results']['analysis_values']['postcal']['standard_2']['img1'][5]
                green_32x_w1_p2 = json_data['test_results']['analysis_values']['postcal']['standard_2']['img2'][3]
                green_32x_w2_p2 = json_data['test_results']['analysis_values']['postcal']['standard_2']['img2'][4]
                green_32x_w3_p2 = json_data['test_results']['analysis_values']['postcal']['standard_2']['img2'][5]
                green_32x_w1_p3 = json_data['test_results']['analysis_values']['postcal']['standard_2']['img3'][3]
                green_32x_w2_p3 = json_data['test_results']['analysis_values']['postcal']['standard_2']['img3'][4]
                green_32x_w3_p3 = json_data['test_results']['analysis_values']['postcal']['standard_2']['img3'][5]

                green_64x_w1 = json_data['test_results']['analysis_values']['postcal']['standard_3']['img1'][3]
                green_64x_w2 = json_data['test_results']['analysis_values']['postcal']['standard_3']['img1'][4]
                green_64x_w3 = json_data['test_results']['analysis_values']['postcal']['standard_3']['img1'][5]
                green_64x_w1_p2 = json_data['test_results']['analysis_values']['postcal']['standard_3']['img2'][3]
                green_64x_w2_p2 = json_data['test_results']['analysis_values']['postcal']['standard_3']['img2'][4]
                green_64x_w3_p2 = json_data['test_results']['analysis_values']['postcal']['standard_3']['img2'][5]
                green_64x_w1_p3 = json_data['test_results']['analysis_values']['postcal']['standard_3']['img3'][3]
                green_64x_w2_p3 = json_data['test_results']['analysis_values']['postcal']['standard_3']['img3'][4]
                green_64x_w3_p3 = json_data['test_results']['analysis_values']['postcal']['standard_3']['img3'][5]

                #Green Standards 4,5,6
                green_16x_w1_std4 = json_data['test_results']['analysis_values']['postcal']['standard_4']['img1'][3]
                green_16x_w2_std4 = json_data['test_results']['analysis_values']['postcal']['standard_4']['img1'][4]
                green_16x_w3_std4 = json_data['test_results']['analysis_values']['postcal']['standard_4']['img1'][5]
                green_16x_w1_p2_std4 = json_data['test_results']['analysis_values']['postcal']['standard_4']['img2'][3]
                green_16x_w2_p2_std4 = json_data['test_results']['analysis_values']['postcal']['standard_4']['img2'][4]
                green_16x_w3_p2_std4 = json_data['test_results']['analysis_values']['postcal']['standard_4']['img2'][5]
                green_16x_w1_p3_std4 = json_data['test_results']['analysis_values']['postcal']['standard_4']['img3'][3]
                green_16x_w2_p3_std4 = json_data['test_results']['analysis_values']['postcal']['standard_4']['img3'][4]
                green_16x_w3_p3_std4 = json_data['test_results']['analysis_values']['postcal']['standard_4']['img3'][5]

                green_32x_w1_std5 = json_data['test_results']['analysis_values']['postcal']['standard_5']['img1'][3]
                green_32x_w2_std5 = json_data['test_results']['analysis_values']['postcal']['standard_5']['img1'][4]
                green_32x_w3_std5 = json_data['test_results']['analysis_values']['postcal']['standard_5']['img1'][5]
                green_32x_w1_p2_std5 = json_data['test_results']['analysis_values']['postcal']['standard_5']['img2'][3]
                green_32x_w2_p2_std5 = json_data['test_results']['analysis_values']['postcal']['standard_5']['img2'][4]
                green_32x_w3_p2_std5 = json_data['test_results']['analysis_values']['postcal']['standard_5']['img2'][5]
                green_32x_w1_p3_std5 = json_data['test_results']['analysis_values']['postcal']['standard_5']['img3'][3]
                green_32x_w2_p3_std5 = json_data['test_results']['analysis_values']['postcal']['standard_5']['img3'][4]
                green_32x_w3_p3_std5 = json_data['test_results']['analysis_values']['postcal']['standard_5']['img3'][5]

                green_64x_w1_std6 = json_data['test_results']['analysis_values']['postcal']['standard_6']['img1'][3]
                green_64x_w2_std6 = json_data['test_results']['analysis_values']['postcal']['standard_6']['img1'][4]
                green_64x_w3_std6 = json_data['test_results']['analysis_values']['postcal']['standard_6']['img1'][5]
                green_64x_w1_p2_std6 = json_data['test_results']['analysis_values']['postcal']['standard_6']['img2'][3]
                green_64x_w2_p2_std6 = json_data['test_results']['analysis_values']['postcal']['standard_6']['img2'][4]
                green_64x_w3_p2_std6 = json_data['test_results']['analysis_values']['postcal']['standard_6']['img2'][5]
                green_64x_w1_p3_std6 = json_data['test_results']['analysis_values']['postcal']['standard_6']['img3'][3]
                green_64x_w2_p3_std6 = json_data['test_results']['analysis_values']['postcal']['standard_6']['img3'][4]
                green_64x_w3_p3_std6 = json_data['test_results']['analysis_values']['postcal']['standard_6']['img3'][5]

                #Buffer Red 7,8,9
                buffer_red_16x_w1 = json_data['test_results']['analysis_values']['postcal']['standard_7']['img1'][0]
                buffer_red_16x_w2 = json_data['test_results']['analysis_values']['postcal']['standard_7']['img1'][1]
                buffer_red_16x_w3 = json_data['test_results']['analysis_values']['postcal']['standard_7']['img1'][2]
                buffer_red_16x_w1_p2 = json_data['test_results']['analysis_values']['postcal']['standard_7']['img2'][0]
                buffer_red_16x_w2_p2 = json_data['test_results']['analysis_values']['postcal']['standard_7']['img2'][1]
                buffer_red_16x_w3_p2 = json_data['test_results']['analysis_values']['postcal']['standard_7']['img2'][2]
                buffer_red_16x_w1_p3 = json_data['test_results']['analysis_values']['postcal']['standard_7']['img3'][0]
                buffer_red_16x_w2_p3 = json_data['test_results']['analysis_values']['postcal']['standard_7']['img3'][1]
                buffer_red_16x_w3_p3 = json_data['test_results']['analysis_values']['postcal']['standard_7']['img3'][2]

                buffer_red_32x_w1 = json_data['test_results']['analysis_values']['postcal']['standard_8']['img1'][0]
                buffer_red_32x_w2 = json_data['test_results']['analysis_values']['postcal']['standard_8']['img1'][1]
                buffer_red_32x_w3 = json_data['test_results']['analysis_values']['postcal']['standard_8']['img1'][2]
                buffer_red_32x_w1_p2 = json_data['test_results']['analysis_values']['postcal']['standard_8']['img2'][0]
                buffer_red_32x_w2_p2 = json_data['test_results']['analysis_values']['postcal']['standard_8']['img2'][1]
                buffer_red_32x_w3_p2 = json_data['test_results']['analysis_values']['postcal']['standard_8']['img2'][2]
                buffer_red_32x_w1_p3 = json_data['test_results']['analysis_values']['postcal']['standard_8']['img3'][0]
                buffer_red_32x_w2_p3 = json_data['test_results']['analysis_values']['postcal']['standard_8']['img3'][1]
                buffer_red_32x_w3_p3 = json_data['test_results']['analysis_values']['postcal']['standard_8']['img3'][2]

                buffer_red_64x_w1 = json_data['test_results']['analysis_values']['postcal']['standard_9']['img1'][0]
                buffer_red_64x_w2 = json_data['test_results']['analysis_values']['postcal']['standard_9']['img1'][1]
                buffer_red_64x_w3 = json_data['test_results']['analysis_values']['postcal']['standard_9']['img1'][2]
                buffer_red_64x_w1_p2 = json_data['test_results']['analysis_values']['postcal']['standard_9']['img2'][0]
                buffer_red_64x_w2_p2 = json_data['test_results']['analysis_values']['postcal']['standard_9']['img2'][1]
                buffer_red_64x_w3_p2 = json_data['test_results']['analysis_values']['postcal']['standard_9']['img2'][2]
                buffer_red_64x_w1_p3 = json_data['test_results']['analysis_values']['postcal']['standard_9']['img3'][0]
                buffer_red_64x_w2_p3 = json_data['test_results']['analysis_values']['postcal']['standard_9']['img3'][1]
                buffer_red_64x_w3_p3 = json_data['test_results']['analysis_values']['postcal']['standard_9']['img3'][2]

                #Buffer Green 7,8,9
                buffer_green_16x_w1 = json_data['test_results']['analysis_values']['postcal']['standard_7']['img1'][3]
                buffer_green_16x_w2 = json_data['test_results']['analysis_values']['postcal']['standard_7']['img1'][4]
                buffer_green_16x_w3 = json_data['test_results']['analysis_values']['postcal']['standard_7']['img1'][5]
                buffer_green_16x_w1_p2 = json_data['test_results']['analysis_values']['postcal']['standard_7']['img2'][3]
                buffer_green_16x_w2_p2 = json_data['test_results']['analysis_values']['postcal']['standard_7']['img2'][4]
                buffer_green_16x_w3_p2 = json_data['test_results']['analysis_values']['postcal']['standard_7']['img2'][5]
                buffer_green_16x_w1_p3 = json_data['test_results']['analysis_values']['postcal']['standard_7']['img3'][3]
                buffer_green_16x_w2_p3 = json_data['test_results']['analysis_values']['postcal']['standard_7']['img3'][4]
                buffer_green_16x_w3_p3 = json_data['test_results']['analysis_values']['postcal']['standard_7']['img3'][5]

                buffer_green_32x_w1 = json_data['test_results']['analysis_values']['postcal']['standard_8']['img1'][3]
                buffer_green_32x_w2 = json_data['test_results']['analysis_values']['postcal']['standard_8']['img1'][4]
                buffer_green_32x_w3 = json_data['test_results']['analysis_values']['postcal']['standard_8']['img1'][5]
                buffer_green_32x_w1_p2 = json_data['test_results']['analysis_values']['postcal']['standard_8']['img2'][3]
                buffer_green_32x_w2_p2 = json_data['test_results']['analysis_values']['postcal']['standard_8']['img2'][4]
                buffer_green_32x_w3_p2 = json_data['test_results']['analysis_values']['postcal']['standard_8']['img2'][5]
                buffer_green_32x_w1_p3 = json_data['test_results']['analysis_values']['postcal']['standard_8']['img3'][3]
                buffer_green_32x_w2_p3 = json_data['test_results']['analysis_values']['postcal']['standard_8']['img3'][4]
                buffer_green_32x_w3_p3 = json_data['test_results']['analysis_values']['postcal']['standard_8']['img3'][5]

                buffer_green_64x_w1 = json_data['test_results']['analysis_values']['postcal']['standard_9']['img1'][3]
                buffer_green_64x_w2 = json_data['test_results']['analysis_values']['postcal']['standard_9']['img1'][4]
                buffer_green_64x_w3 = json_data['test_results']['analysis_values']['postcal']['standard_9']['img1'][5]
                buffer_green_64x_w1_p2 = json_data['test_results']['analysis_values']['postcal']['standard_9']['img2'][3]
                buffer_green_64x_w2_p2 = json_data['test_results']['analysis_values']['postcal']['standard_9']['img2'][4]
                buffer_green_64x_w3_p2 = json_data['test_results']['analysis_values']['postcal']['standard_9']['img2'][5]
                buffer_green_64x_w1_p3 = json_data['test_results']['analysis_values']['postcal']['standard_9']['img3'][3]
                buffer_green_64x_w2_p3 = json_data['test_results']['analysis_values']['postcal']['standard_9']['img3'][4]
                buffer_green_64x_w3_p3 = json_data['test_results']['analysis_values']['postcal']['standard_9']['img3'][5]

                #Target Green Standards 1,2,3
                target_green_std123_w1 = json_data['test_results']['ref_values']['green'][0]
                target_green_std123_w2 = json_data['test_results']['ref_values']['green'][1]
                target_green_std123_w3 = json_data['test_results']['ref_values']['green'][2]

                #Target Red Standards 4,5,6
                target_red_std123_w1 = json_data['test_results']['ref_values']['red'][0]
                target_red_std123_w2 = json_data['test_results']['ref_values']['red'][1]
                target_red_std123_w3 = json_data['test_results']['ref_values']['red'][2]

                #Normalizers red
                r_norm_w1 = json_data['test_results']['spec_values']['data']['normalizers']['r_norm'][0]
                r_norm_w2 = json_data['test_results']['spec_values']['data']['normalizers']['r_norm'][1]
                r_norm_w3 = json_data['test_results']['spec_values']['data']['normalizers']['r_norm'][2]

                #Normalizers green
                g_norm_w1 = json_data['test_results']['spec_values']['data']['normalizers']['g_norm'][0]
                g_norm_w2 = json_data['test_results']['spec_values']['data']['normalizers']['g_norm'][1]
                g_norm_w3 = json_data['test_results']['spec_values']['data']['normalizers']['g_norm'][2]

                #Red Linearity
                l_red_w1 = json_data['test_results']['spec_values']['spec_results']['well_1']['lin_red']
                l_red_w2 = json_data['test_results']['spec_values']['spec_results']['well_2']['lin_red']
                l_red_w3 = json_data['test_results']['spec_values']['spec_results']['well_3']['lin_red']

                #Green Linearity
                l_green_w1 = json_data['test_results']['spec_values']['spec_results']['well_1']['lin_green']
                l_green_w2 = json_data['test_results']['spec_values']['spec_results']['well_2']['lin_green']
                l_green_w3 = json_data['test_results']['spec_values']['spec_results']['well_3']['lin_green']

        dummy = 0 #dummy variable for printing the photo ID and Well Number
        output_df.loc[row_index] = [hostname_data, ref_unit_data, batch_code_data, date_time_data, int(dummy+1), l_red_w1, "Pass" if l_red_w1 >=0.9950 else "Fail", l_green_w1, "Pass" if l_green_w1 >=0.9950 else "Fail", r_norm_w1, g_norm_w1, well_1_center, red_16x_w1, red_32x_w1, red_64x_w1, red_16x_w1_std4, "Pass" if 24525<=red_16x_w1_std4<=33579 else "Fail", red_32x_w1_std5, "Pass" if 9762<=red_32x_w1_std5<=17118 else "Fail", red_64x_w1_std6, "Pass" if 3505<=red_64x_w1_std6<=7885 else "Fail", green_16x_w1, "Pass" if 20564<=green_16x_w1<=37820 else "Fail", green_32x_w1, "Pass" if 9578<=green_32x_w1<=18860 else "Fail", green_64x_w1, "Pass" if 3752<=green_64x_w1<=9080 else "Fail", green_16x_w1_std4, green_32x_w1_std5, green_64x_w1_std6, buffer_red_16x_w1, buffer_red_32x_w1, buffer_red_64x_w1, buffer_green_16x_w1, buffer_green_32x_w1, buffer_green_64x_w1, target_green_std123_w1, target_red_std123_w1, int(dummy+1)]        
        output_df.loc[row_index+1] = [hostname_data, ref_unit_data, batch_code_data, date_time_data, int(dummy+2), l_red_w2, "Pass" if l_red_w2 >=0.9950 else "Fail", l_green_w2, "Pass" if l_green_w2 >=0.9950 else "Fail", r_norm_w2, g_norm_w2, well_2_center, red_16x_w2, red_32x_w2, red_64x_w2, red_16x_w2_std4, "Pass" if 23185<=red_16x_w2_std4<=34885 else "Fail", red_32x_w2_std5, "Pass" if 10224<=red_32x_w2_std5<=17406 else "Fail", red_64x_w2_std6, "Pass" if 3619<=red_64x_w2_std6<=8053 else "Fail", green_16x_w2, "Pass" if 22076<=green_16x_w2<=36890 else "Fail", green_32x_w2, "Pass" if 10489<=green_32x_w2<=18667 else "Fail", green_64x_w2, "Pass" if 3909<=green_64x_w2<=9105 else "Fail", green_16x_w2_std4, green_32x_w2_std5, green_64x_w2_std6, buffer_red_16x_w2, buffer_red_32x_w2, buffer_red_64x_w2, buffer_green_16x_w2, buffer_green_32x_w2, buffer_green_64x_w2, target_green_std123_w2, target_red_std123_w2, int(dummy+1)]
        output_df.loc[row_index+2] = [hostname_data, ref_unit_data, batch_code_data, date_time_data, int(dummy+3), l_red_w3, "Pass" if l_red_w3 >=0.9950 else "Fail", l_green_w3, "Pass" if l_green_w3 >=0.9950 else "Fail", r_norm_w3, g_norm_w3, well_3_center, red_16x_w3, red_32x_w3, red_64x_w3, red_16x_w3_std4, "Pass" if 25189<=red_16x_w3_std4<=32761 else "Fail", red_32x_w3_std5, "Pass" if 9865<=red_32x_w3_std5<=17017 else "Fail", red_64x_w3_std6, "Pass" if 3365<=red_64x_w3_std6<=7853 else "Fail", green_16x_w3, "Pass" if 21750<=green_16x_w3<=35256 else "Fail", green_32x_w3, "Pass" if 9980<=green_32x_w3<=18134 else "Fail", green_64x_w3, "Pass" if 3543<=green_64x_w3<=8811 else "Fail", green_16x_w3_std4, green_32x_w3_std5, green_64x_w3_std6, buffer_red_16x_w3, buffer_red_32x_w3, buffer_red_64x_w3, buffer_green_16x_w3, buffer_green_32x_w3, buffer_green_64x_w3, target_green_std123_w3, target_red_std123_w3, int(dummy+1)]

        output_df.loc[row_index+3] = [hostname_data, ref_unit_data, batch_code_data, date_time_data, int(dummy+1), l_red_w1, "Pass" if l_red_w1 >=0.9950 else "Fail", l_green_w1, "Pass" if l_green_w1 >=0.9950 else "Fail", r_norm_w1, g_norm_w1, well_1_center, red_16x_w1_p2, red_32x_w1_p2, red_64x_w1_p2, red_16x_w1_p2_std4, "Pass" if 24525<=red_16x_w1_p2_std4<=33579 else "Fail", red_32x_w1_p2_std5, "Pass" if 9762<=red_32x_w1_p2_std5<=17118 else "Fail", red_64x_w1_p2_std6, "Pass" if 3505<=red_64x_w1_p2_std6<=7885 else "Fail", green_16x_w1_p2, "Pass" if 20564<=green_16x_w1_p2<=37820 else "Fail", green_32x_w1_p2, "Pass" if 9578<=green_32x_w1_p2<=18860 else "Fail", green_64x_w1_p2, "Pass" if 3752<=green_64x_w1_p2<=9080 else "Fail", green_16x_w1_p2_std4, green_32x_w1_p2_std5, green_64x_w1_p2_std6, buffer_red_16x_w1_p2, buffer_red_32x_w1_p2, buffer_red_64x_w1_p2, buffer_green_16x_w1_p2, buffer_green_32x_w1_p2, buffer_green_64x_w1_p2, target_green_std123_w1, target_red_std123_w1, int(dummy+2)]
        output_df.loc[row_index+4] = [hostname_data, ref_unit_data, batch_code_data, date_time_data, int(dummy+2), l_red_w2, "Pass" if l_red_w2 >=0.9950 else "Fail", l_green_w2, "Pass" if l_green_w2 >=0.9950 else "Fail", r_norm_w2, g_norm_w2, well_2_center, red_16x_w2_p2, red_32x_w2_p2, red_64x_w2_p2, red_16x_w2_p2_std4, "Pass" if 23185<=red_16x_w2_p2_std4<=34885 else "Fail", red_32x_w2_p2_std5, "Pass" if 10224<=red_32x_w2_p2_std5<=17406 else "Fail", red_64x_w2_p2_std6, "Pass" if 3619<=red_64x_w2_p2_std6<=8053 else "Fail", green_16x_w2_p2, "Pass" if 22076<=green_16x_w2_p2<=36890 else "Fail", green_32x_w2_p2, "Pass" if 10489<=green_32x_w2_p2<=18667 else "Fail", green_64x_w2_p2, "Pass" if 3909<=green_64x_w2_p2<=9105 else "Fail", green_16x_w2_p2_std4, green_32x_w2_p2_std5, green_64x_w2_p2_std6, buffer_red_16x_w2_p2, buffer_red_32x_w2_p2, buffer_red_64x_w2_p2, buffer_green_16x_w2_p2, buffer_green_32x_w2_p2, buffer_green_64x_w2_p2, target_green_std123_w2, target_red_std123_w2, int(dummy+2)]
        output_df.loc[row_index+5] = [hostname_data, ref_unit_data, batch_code_data, date_time_data, int(dummy+3), l_red_w3, "Pass" if l_red_w3 >=0.9950 else "Fail", l_green_w3, "Pass" if l_green_w3 >=0.9950 else "Fail", r_norm_w3, g_norm_w3, well_3_center, red_16x_w3_p2, red_32x_w3_p2, red_64x_w3_p2, red_16x_w3_p2_std4, "Pass" if 25189<=red_16x_w3_p2_std4<=32761 else "Fail", red_32x_w3_p2_std5, "Pass" if 9865<=red_32x_w3_p2_std5<=17017 else "Fail", red_64x_w3_p2_std6, "Pass" if 3365<=red_64x_w3_p2_std6<=7853 else "Fail", green_16x_w3_p2, "Pass" if 21750<=green_16x_w3_p2<=35256 else "Fail", green_32x_w3_p2, "Pass" if 9980<=green_32x_w3_p2<=18134 else "Fail", green_64x_w3_p2, "Pass" if 3543<=green_64x_w3_p2<=8811 else "Fail", green_16x_w3_p2_std4, green_32x_w3_p2_std5, green_64x_w3_p2_std6, buffer_red_16x_w3_p2, buffer_red_32x_w3_p2, buffer_red_64x_w3_p2, buffer_green_16x_w3_p2, buffer_green_32x_w3_p2, buffer_green_64x_w3_p2, target_green_std123_w3, target_red_std123_w3, int(dummy+2)]

        output_df.loc[row_index+6] = [hostname_data, ref_unit_data, batch_code_data, date_time_data, int(dummy+1), l_red_w1, "Pass" if l_red_w1 >=0.9950 else "Fail", l_green_w1, "Pass" if l_green_w1 >=0.9950 else "Fail", r_norm_w1, g_norm_w1, well_1_center, red_16x_w1_p3, red_32x_w1_p3, red_64x_w1_p3, red_16x_w1_p3_std4, "Pass" if 24525<=red_16x_w1_p3_std4<=33579 else "Fail", red_32x_w1_p3_std5, "Pass" if 9762<=red_32x_w1_p3_std5<=17118 else "Fail", red_64x_w1_p3_std6, "Pass" if 3505<=red_64x_w1_p3_std6<=7885 else "Fail", green_16x_w1_p3, "Pass" if 20564<=green_16x_w1_p3<=37820 else "Fail", green_32x_w1_p3, "Pass" if 9578<=green_32x_w1_p3<=18860 else "Fail", green_64x_w1_p3, "Pass" if 3752<=green_64x_w1_p3<=9080 else "Fail", green_16x_w1_p3_std4, green_32x_w1_p3_std5, green_64x_w1_p3_std6, buffer_red_16x_w1_p3, buffer_red_32x_w1_p3, buffer_red_64x_w1_p3, buffer_green_16x_w1_p3, buffer_green_32x_w1_p3, buffer_green_64x_w1_p3, target_green_std123_w1, target_red_std123_w1, int(dummy+3)]
        output_df.loc[row_index+7] = [hostname_data, ref_unit_data, batch_code_data, date_time_data, int(dummy+2), l_red_w2, "Pass" if l_red_w2 >=0.9950 else "Fail", l_green_w2, "Pass" if l_green_w2 >=0.9950 else "Fail", r_norm_w2, g_norm_w2, well_2_center, red_16x_w2_p3, red_32x_w2_p3, red_64x_w2_p3, red_16x_w2_p3_std4, "Pass" if 23185<=red_16x_w2_p3_std4<=34885 else "Fail", red_32x_w2_p3_std5, "Pass" if 10224<=red_32x_w2_p3_std5<=17406 else "Fail", red_64x_w2_p3_std6, "Pass" if 3619<=red_64x_w2_p3_std6<=8053 else "Fail", green_16x_w2_p3, "Pass" if 22076<=green_16x_w2_p3<=36890 else "Fail", green_32x_w2_p3, "Pass" if 10489<=green_32x_w2_p3<=18667 else "Fail", green_64x_w2_p3, "Pass" if 3909<=green_64x_w2_p3<=9105 else "Fail", green_16x_w2_p3_std4, green_32x_w2_p3_std5, green_64x_w2_p3_std6, buffer_red_16x_w2_p3, buffer_red_32x_w2_p3, buffer_red_64x_w2_p3, buffer_green_16x_w2_p3, buffer_green_32x_w2_p3, buffer_green_64x_w2_p3, target_green_std123_w2, target_red_std123_w2, int(dummy+3)]
        output_df.loc[row_index+8] = [hostname_data, ref_unit_data, batch_code_data, date_time_data, int(dummy+3), l_red_w3, "Pass" if l_red_w3 >=0.9950 else "Fail", l_green_w3, "Pass" if l_green_w3 >=0.9950 else "Fail", r_norm_w3, g_norm_w3, well_3_center, red_16x_w3_p3, red_32x_w3_p3, red_64x_w3_p3, red_16x_w3_p3_std4, "Pass" if 25189<=red_16x_w3_p3_std4<=32761 else "Fail", red_32x_w3_p3_std5, "Pass" if 9865<=red_32x_w3_p3_std5<=17017 else "Fail", red_64x_w3_p3_std6, "Pass" if 3365<=red_64x_w3_p3_std6<=7853 else "Fail", green_16x_w3_p3, "Pass" if 21750<=green_16x_w3_p3<=35256 else "Fail", green_32x_w3_p3, "Pass" if 9980<=green_32x_w3_p3<=18134 else "Fail", green_64x_w3_p3, "Pass" if 3543<=green_64x_w3_p3<=8811 else "Fail", green_16x_w3_p3_std4, green_32x_w3_p3_std5, green_64x_w3_p3_std6, buffer_red_16x_w3_p3, buffer_red_32x_w3_p3, buffer_red_64x_w3_p3, buffer_green_16x_w3_p3, buffer_green_32x_w3_p3, buffer_green_64x_w3_p3, target_green_std123_w3, target_red_std123_w3, int(dummy+3)]

def flatten_dict(d, parent_key='', sep='_'):
    items = []
    for k, v in d.items():
        new_key = parent_key + sep + k if parent_key else k
        if isinstance(v, dict):
            items.extend(flatten_dict(v, new_key, sep=sep).items())
        elif isinstance(v, list):
            for i, val in enumerate(v, 1):
                items.append((f"{new_key}_{i}", val))
        else:
            items.append((new_key, v))
    return dict(items)        

def process_folder(folder_path):
    tar_files = glob.glob(os.path.join(folder_path, "**/*.tar.gz"), recursive=True)
    output_df = pd.DataFrame(columns=["Cube-HN", "Ref-Cube-HN", "Batchcode", "Date-Time", "Well-Number", "Red-Linearity", "RL-Control", "Green-Linearity", "GL-Control", "Red-Norm", "Green-Norm", "Well Center Coordinatess", "Red-16X-Standard-1", "Red-32X-Standard-2", "Red-64X-Standard-3", "Red-16X-Standard-4", "R-16X-Control", "Red-32X-Standard-5", "R-32X-Control", "Red-64X-Standard-6", "R-64X-Control", "Green-16X-Standard-1", "G-16X-Control", "Green-32X-Standard-2", "G-32X-Control", "Green-64X-Standard-3", "G-64X-Control", "Green-16X-Standard-4", "Green-32X-Standard-5", "Green-64X-Standard-6", "Buffer-Red-Standard-7", "Buffer-Red-Standard-8", "Buffer-Red-Standard-9", "Buffer-Green-Standard-7", "Buffer-Green-Standard-8", "Buffer-Green-Standard-9", "Target(Green-Std-1,2,3)", "Target(Red-Std-4,5,6)", "Photo-ID"])
    workbook_path = "DEV-00196 Attachment #1 - Single Cube Calibration Data Evaluation.xlsx"
    workbook = openpyxl.load_workbook(workbook_path)
    for i, tar_file in enumerate(tar_files):
        print(tar_file)
        process_tar_file(tar_file, output_df, i*9, workbook)
        #print(output_df)
        d=i*9
        hn = output_df["Cube-HN"][d]
        dt = output_df["Date-Time"][d]
        print(hn)
        print(dt)

        #home_path = os.environ["HOMEPATH"] + "\\Desktop\\"
        if os.path.isdir(os.path.join(os.environ["HOMEPATH"], "Desktop", "CubeData")):
            output_file = os.path.join(os.environ["HOMEPATH"], "Desktop", "CubeData\\{}-{}.xlsx".format(hn, dt))
        else:
            os.mkdir(os.path.join(os.environ["HOMEPATH"], "Desktop", "CubeData"))
            output_file = os.path.join(os.environ["HOMEPATH"], "Desktop", "CubeData\\{}-{}.xlsx".format(hn, dt))
        #output_file = "C:\\Users\\UKumar\\Desktop\\CubeData\\{}-{}.xlsx".format(hn, dt)

        workbook.save(output_file)
        for row in workbook["optical_store"].iter_rows():
            for cell in row:
                cell.value = None
        for row in workbook["optical_targets"].iter_rows():
            for cell in row:
                cell.value = None
        # output_file = "C:\\Users\\UKumar\\Desktop\\MES-Study\\{}-{}.xlsx".format(hn, dt)
        # output_df.to_excel(output_file, index=False)
        # output_df = output_df.drop(output_df.index[:9])
    return output_df

def browse_folder():
    folder_path = filedialog.askdirectory()
    folder_entry.delete(0, tk.END)
    folder_entry.insert(0, folder_path)

def run_script():
    folder_path = folder_entry.get()
    df = process_folder(folder_path)

'''
if __name__ == "__main__":
    folder_path = input("Enter the folder path: ")
    df = process_folder(folder_path)
    #print(df)
    #output_file = "C:\\Users\\UKumar\\Desktop\\derp.xlsx"
    #df.to_excel(output_file, index=False)
'''

# Create the main tkinter window
window = tk.Tk()
window.title("MES Extractor")

# Folder Selection
folder_label = tk.Label(window, text="Folder Path:")
folder_label.pack()

folder_entry = tk.Entry(window, width=50)
folder_entry.pack()

browse_button = tk.Button(window, text="Browse", command=browse_folder)
browse_button.pack()

# Run Script Button
run_button = tk.Button(window, text="Run Script", command=run_script)
run_button.pack()

# Start the tkinter event loop
window.mainloop()